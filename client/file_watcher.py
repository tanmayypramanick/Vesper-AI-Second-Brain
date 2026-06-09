#!/usr/bin/env python3
"""
VESPER MAC AGENT - Production v6
══════════════════════════════════
Key improvements:
- Throttled file sending (no CPU spike)
- No OCR on Mac (too heavy, done server-side later)
- Batched state saves (every 50 files)
- Videos queued for overnight batch only
- Smart rate limiting
"""

import os
import sys
import json
import time
import shutil
import hashlib
import logging
import threading
import subprocess
import http.server
import socketserver
from datetime import datetime, timedelta
from pathlib import Path

import requests
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# ══════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════

VESPER_IP         = "10.0.0.120"
VESPER_IP_TAILSCALE = "100.123.15.32"   # always reachable via Tailscale VPN
VESPER_PORT       = 5000
VESPER_URL        = f"http://{VESPER_IP}:{VESPER_PORT}/ingest"
VESPER_URL_BACKUP = f"http://{VESPER_IP_TAILSCALE}:{VESPER_PORT}/ingest"
COMMAND_PORT      = 5001

DEBOUNCE_SECS   = 5.0
REQUEST_TIMEOUT = 120   # 120s — server embeds on slow CPU; server should go async (see prompt)
CHUNK_SIZE      = 800   # chars per chunk
MAX_CHUNKS      = 6     # max chunks per file
MAX_FILE_MB     = 20    # skip files larger than this
MAX_JSON_KB     = 100   # raised from 30KB — catches more data files
MAX_PDF_PAGES   = 20    # raised from 8 — full coverage of longer docs

# Throttle between file sends — prevents server CPU spike
# 0.5s = 2 files/second max. Adjust down to 0.2 if server handles it
SEND_DELAY = 0.5

# Save state every N files (not every file)
STATE_SAVE_INTERVAL = 50

AGENT_DIR  = os.path.expanduser("~/vesper_agent")
QUEUE_FILE = f"{AGENT_DIR}/queue/pending.json"
STATE_FILE = f"{AGENT_DIR}/indexed_state.json"
LOG_FILE   = f"{AGENT_DIR}/logs/watcher.log"

WATCH_DIRS = [d for d in [
    os.path.expanduser("~/Documents"),
    os.path.expanduser("~/Desktop"),
    os.path.expanduser("~/Downloads"),
    os.path.expanduser("~/Developer"),
    os.path.expanduser("~/Projects"),
    os.path.expanduser("~/Code"),
    os.path.expanduser("~/repos"),
    os.path.expanduser("~/obsidian"),
    os.path.expanduser("~/Obsidian"),
    os.path.expanduser("~/Movies"),
    os.path.expanduser("~/Pictures"),
    os.path.expanduser("~/Music"),
] if os.path.exists(d)]

TEXT_EXTENSIONS = {
    # Documents
    ".txt", ".md", ".markdown", ".rtf", ".org",
    ".pdf", ".docx", ".doc", ".pages", ".odt",
    ".csv", ".xlsx", ".xls", ".numbers",
    # Web / UI
    ".py", ".swift", ".js", ".ts", ".tsx", ".jsx", ".go",
    ".html", ".htm", ".css", ".scss", ".sass",
    ".sh", ".bash", ".zsh", ".fish",
    # Data / Config
    ".yaml", ".yml", ".toml", ".json", ".jsonl", ".xml",
    ".env", ".ini", ".cfg", ".conf",
    # Docs / Academic
    ".rst", ".tex", ".sql", ".r", ".ipynb",
    # Systems / Misc
    ".java", ".kt", ".cpp", ".c", ".h", ".hpp",
    ".rb", ".php", ".rs", ".dart", ".lua", ".scala",
    ".svg",   # SVG is XML text — readable
    ".plist", # Apple property list — readable XML
}

# Images: store metadata only (no OCR on Mac)
# OCR is too CPU-heavy for a background agent
IMAGE_EXTENSIONS = {
    ".jpg", ".jpeg", ".png", ".gif",
    ".webp", ".bmp", ".tiff", ".heic",
    ".heif", ".avif",
}

# Videos: deferred to overnight batch
VIDEO_EXTENSIONS = {
    ".mp4", ".mov", ".avi", ".mkv",
    ".m4v", ".wmv", ".flv", ".webm",
}

ALL_SUPPORTED = TEXT_EXTENSIONS | IMAGE_EXTENSIONS | VIDEO_EXTENSIONS

SKIP_FILENAMES = {
    "package-lock.json", "yarn.lock",
    "poetry.lock", "Pipfile.lock",
    "Cargo.lock", ".DS_Store", "Thumbs.db",
    "tsconfig.json", "next-env.d.ts", ".eslintcache",
    "workbench.desktop.main.js", "workbench.desktop.main.css",
    "extensionHostProcess.js", "extensionHostWorkerMain.js",
    "notebookSimpleWorkerMain.js", "languageDetectionSimpleWorkerMain.js",
    "textMateTokenizationWorker.workerMain.js", "outputLinkComputerMain.js",
    "localFileSearchMain.js", "editorSimpleWorkerMain.js",
    "processExplorerMain.js", "processExplorerMain.css",
    "telemetryApp.js", "sharedProcessMain.js", "ptyHostMain.js",
    "watcherMain.js", "cliProcessMain.js", "diagnosticTool.js",
    "diagnosticTool.css", "profileAnalysisWorkerMain.js",
    "renameWorker.js", "watchdog.js", "bootloader.js",
    "cellAttachmentRenderer.js", "preload.js", "bootstrap-fork.js",
    "cli.js", "extension.js", "service-worker.js",
    "ThirdPartyNotices.txt", "LICENSES.chromium.html",
    "workbench.html", "workbench.js",
}

SKIP_DIRS = {
    "node_modules", ".git", "__pycache__",
    ".next", ".nuxt", "dist", "build",
    ".venv", "venv", "env", "Caches",
    ".Trash", "tmp", "temp", "Library",
    "vendor", ".yarn", ".cache", ".pnpm-store",
    "PhotoLibrary", "Visual Studio Code.app",
    "VSCode", "vscode", ".vscode", "Code.app",
    "Electron.app", "resources", "extensions",
    # Apple internal — never useful to index
    "spotlightV3", "Spotlight", "CoreData",
    "NSFileProtectionComplete", "NSFileProtectionCompleteUntilFirstUserAuthentication",
    "index.spotlightV3",
}

# ══════════════════════════════════════════
# LOGGING — reduced verbosity
# ══════════════════════════════════════════

os.makedirs(f"{AGENT_DIR}/logs",  exist_ok=True)
os.makedirs(f"{AGENT_DIR}/queue", exist_ok=True)

# LaunchAgent already captures stdout → watcher.log via StandardOutPath.
# Only use FileHandler to avoid every line being written twice.
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE),
    ]
)
log = logging.getLogger("vesper.agent")

# ══════════════════════════════════════════
# PROGRESS TRACKER
# ══════════════════════════════════════════

class ProgressTracker:
    def __init__(self):
        self.total = self.processed = 0
        self.synced = self.skipped = self.failed = 0
        self.start_time = None
        self.lock = threading.Lock()

    def start(self, total: int):
        with self.lock:
            self.total      = total
            self.processed  = 0
            self.synced     = 0
            self.skipped    = 0
            self.failed     = 0
            self.start_time = datetime.now()

    def update(self, synced=0, skipped=0, failed=0):
        with self.lock:
            self.processed += 1
            self.synced    += synced
            self.skipped   += skipped
            self.failed    += failed
            if self.processed % 25 == 0 or \
               self.processed == self.total:
                self._print()

    def _print(self):
        if not self.start_time or self.total == 0:
            return
        pct     = (self.processed / self.total) * 100
        elapsed = (datetime.now() - self.start_time).total_seconds()
        if self.processed > 0:
            rate    = elapsed / self.processed
            remain  = (self.total - self.processed) * rate
            eta     = str(timedelta(seconds=int(remain)))
        else:
            eta = "calculating..."
        log.info(
            f"📊 {self.processed}/{self.total} ({pct:.0f}%) | "
            f"✅{self.synced} ⏭️{self.skipped} ❌{self.failed} | "
            f"ETA: {eta}"
        )

    def done(self):
        with self.lock:
            elapsed = (datetime.now() - self.start_time
                      ).total_seconds() if self.start_time else 0
            log.info(
                f"\n{'='*50}\n✅ SCAN COMPLETE\n"
                f"   Synced:  {self.synced}\n"
                f"   Skipped: {self.skipped}\n"
                f"   Failed:  {self.failed}\n"
                f"   Time:    {str(timedelta(seconds=int(elapsed)))}\n"
                f"{'='*50}"
            )

progress = ProgressTracker()

# ══════════════════════════════════════════
# CONNECTIVITY
# ══════════════════════════════════════════

def can_reach_vesper() -> bool:
    """Try local IP first (fast LAN), fall back to Tailscale (always works)."""
    for url in [
        f"http://{VESPER_IP}:{VESPER_PORT}/health",
        f"http://{VESPER_IP_TAILSCALE}:{VESPER_PORT}/health",
    ]:
        try:
            r = requests.get(url, timeout=3)
            if r.status_code == 200:
                return True
        except Exception:
            continue
    return False

def is_mac_sleeping() -> bool:
    try:
        result = subprocess.run(
            ["ioreg", "-n", "IODisplayWrangler"],
            capture_output=True, text=True, timeout=3
        )
        return "DevicePowerState=0" in result.stdout
    except Exception:
        return False

# ══════════════════════════════════════════
# STATE (thread-safe, batched saves)
# ══════════════════════════════════════════

_state_lock    = threading.Lock()
_queue_lock    = threading.Lock()
_save_counter  = 0

def load_state() -> dict:
    with _state_lock:
        try:
            if os.path.exists(STATE_FILE):
                with open(STATE_FILE) as f:
                    return json.load(f)
        except Exception:
            pass
        return {}

def save_state(state: dict, force: bool = False) -> None:
    global _save_counter
    _save_counter += 1
    if not force and _save_counter % STATE_SAVE_INTERVAL != 0:
        return
    try:
        with open(STATE_FILE, "w") as f:
            json.dump(state, f)
    except Exception as e:
        log.error(f"State save: {e}")

def compute_hash(filepath: str) -> str | None:
    try:
        h = hashlib.md5()
        with open(filepath, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
        return h.hexdigest()
    except Exception:
        return None

# ══════════════════════════════════════════
# QUEUE
# ══════════════════════════════════════════

def load_queue() -> list:
    with _queue_lock:
        try:
            if os.path.exists(QUEUE_FILE):
                with open(QUEUE_FILE) as f:
                    return json.load(f)
        except Exception:
            pass
        return []

def save_queue(items: list) -> None:
    with _queue_lock:
        try:
            with open(QUEUE_FILE, "w") as f:
                json.dump(items, f)
        except Exception as e:
            log.error(f"Queue save: {e}")

def add_to_queue(filepath: str, event_type: str) -> None:
    items = load_queue()
    items = [i for i in items if i.get("filepath") != filepath]
    items.append({
        "filepath":   filepath,
        "event_type": event_type,
        "queued_at":  datetime.now().isoformat()
    })
    save_queue(items)

def remove_from_queue(filepath: str) -> None:
    items = [i for i in load_queue() if i.get("filepath") != filepath]
    save_queue(items)

# ══════════════════════════════════════════
# FILE FILTERING
# ══════════════════════════════════════════

def should_index(filepath: str) -> bool:
    try:
        p   = Path(filepath)
        ext = p.suffix.lower()

        if p.name in SKIP_FILENAMES:
            return False
        if ext not in ALL_SUPPORTED:
            return False
        for part in p.parts:
            if part in SKIP_DIRS:
                return False
            if part.startswith("."):
                return False
        size = p.stat().st_size
        if size == 0:
            return False
        if ext == ".json" and size > MAX_JSON_KB * 1024:
            return False
        if size > MAX_FILE_MB * 1024 * 1024:
            return False
        return True
    except Exception:
        return False

# ══════════════════════════════════════════
# TEXT EXTRACTION
# ══════════════════════════════════════════

def extract_text(filepath: str) -> str | None:
    ext = Path(filepath).suffix.lower()
    try:
        # Any text-readable extension — read directly as UTF-8
        if ext in {
            ".txt", ".md", ".markdown", ".org",
            ".rst", ".tex", ".py", ".swift",
            ".js", ".ts", ".tsx", ".jsx", ".go",
            ".html", ".htm", ".css", ".scss", ".sass",
            ".sh", ".bash", ".zsh", ".fish",
            ".yaml", ".yml", ".toml", ".json", ".jsonl",
            ".xml", ".svg", ".plist",
            ".env", ".ini", ".cfg", ".conf",
            ".sql", ".r", ".ipynb",
            ".java", ".kt", ".cpp", ".c", ".h", ".hpp",
            ".rb", ".php", ".rs", ".dart", ".lua", ".scala",
            ".rtf",
        }:
            with open(filepath, "r", errors="ignore") as f:
                return f.read()

        elif ext == ".pdf":
            from pypdf import PdfReader
            reader = PdfReader(filepath)
            pages  = []
            for i, page in enumerate(reader.pages):
                if i >= MAX_PDF_PAGES:
                    break
                try:
                    text = page.extract_text()
                    if text and text.strip():
                        pages.append(f"[Page {i+1}]\n{text.strip()}")
                except Exception:
                    continue
            return "\n\n".join(pages) or None

        elif ext in {".docx", ".doc"}:
            from docx import Document
            doc  = Document(filepath)
            return "\n".join([
                p.text for p in doc.paragraphs if p.text.strip()
            ]) or None

        elif ext in {".xlsx", ".xls"}:
            import openpyxl
            wb = openpyxl.load_workbook(
                filepath, read_only=True, data_only=True
            )
            output = []
            for name in wb.sheetnames:
                ws   = wb[name]
                rows = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i > 150:
                        break
                    row_str = " | ".join([
                        str(c) for c in row if c is not None
                    ])
                    if row_str.strip():
                        rows.append(row_str)
                if rows:
                    output.append(f"[Sheet: {name}]\n" + "\n".join(rows))
            return "\n\n".join(output) or None

        elif ext == ".csv":
            lines = []
            with open(filepath, "r", errors="ignore") as f:
                for i, line in enumerate(f):
                    if i >= 150:
                        break
                    if line.strip():
                        lines.append(line.strip())
            return "\n".join(lines) or None

    except Exception as e:
        log.warning(f"Extract {Path(filepath).name}: {e}")
    return None

def extract_image_metadata(filepath: str) -> str | None:
    """
    Store image metadata only — no OCR on Mac.
    OCR is too CPU-heavy for a background agent.
    Vesper knows the image exists and can describe it
    via llava when queried directly.
    """
    p    = Path(filepath)
    try:
        size = p.stat().st_size
        return (
            f"[Image: {p.name}] "
            f"Location: {str(p.parent)} | "
            f"Size: {size//1024}KB | "
            f"Type: {p.suffix.upper()}"
        )
    except Exception:
        return None

def extract_content(filepath: str) -> str | None:
    ext = Path(filepath).suffix.lower()
    if ext in TEXT_EXTENSIONS:
        return extract_text(filepath)
    elif ext in IMAGE_EXTENSIONS:
        return extract_image_metadata(filepath)
    elif ext in VIDEO_EXTENSIONS:
        # Videos: metadata only, transcription done overnight
        p = Path(filepath)
        try:
            size = p.stat().st_size
            return (
                f"[Video: {p.name}] "
                f"Location: {str(p.parent)} | "
                f"Size: {size // (1024*1024)}MB"
            )
        except Exception:
            return None
    return None

# ══════════════════════════════════════════
# CHUNKING
# ══════════════════════════════════════════

def chunk_text(text: str) -> list[str]:
    if not text or len(text.strip()) < 20:
        return []
    text = text.strip()
    if len(text) <= CHUNK_SIZE:
        return [text]

    paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
    chunks  = []
    current = ""

    for para in paragraphs:
        if len(para) > CHUNK_SIZE:
            if current:
                chunks.append(current.strip())
                current = ""
            for i in range(0, len(para), CHUNK_SIZE - 50):
                sub = para[i:i+CHUNK_SIZE-50].strip()
                if sub:
                    chunks.append(sub)
            continue
        if len(current) + len(para) > CHUNK_SIZE:
            if current:
                chunks.append(current.strip())
            current = para
        else:
            current = current + "\n\n" + para if current else para

    if current.strip():
        chunks.append(current.strip())

    return chunks[:MAX_CHUNKS]

# ══════════════════════════════════════════
# SYNC ENGINE — with throttling
# ══════════════════════════════════════════

def sync_file(filepath: str, event_type: str, state: dict) -> bool:
    if event_type == "deleted":
        _notify_deletion(filepath, state)
        return True

    if not os.path.exists(filepath):
        return False

    content = extract_content(filepath)
    if not content or len(content.strip()) < 20:
        return False

    chunks = chunk_text(content)
    if not chunks:
        return False

    p = Path(filepath)

    payload = {
        "filepath":     filepath,
        "filename":     p.name,
        "folder":       str(p.parent),
        "event_type":   event_type,
        "timestamp":    datetime.now().isoformat(),
        "chunks":       chunks,
        "total_chunks": len(chunks),
        "file_type":    p.suffix.lower(),
    }

    # Try local LAN first (fast), fall back to Tailscale (always reachable)
    last_exc = None
    for ingest_url in [VESPER_URL, VESPER_URL_BACKUP]:
        try:
            r = requests.post(ingest_url, json=payload, timeout=REQUEST_TIMEOUT)
            if r.status_code in (200, 202):
                h = compute_hash(filepath)
                if h:
                    with _state_lock:
                        state[filepath] = {
                            "hash":      h,
                            "synced_at": datetime.now().isoformat()
                        }
                        save_state(state)
                log.info(f"✅ {p.name} → {len(chunks)} chunks")
                return True
            log.error(f"❌ {r.status_code}: {p.name}")
            return False
        except requests.exceptions.Timeout:
            log.warning(f"⏱️  Timeout: {p.name} → queued")
            add_to_queue(filepath, event_type)
            return False
        except requests.exceptions.ConnectionError as e:
            last_exc = e
            continue  # try backup URL
        except Exception as e:
            log.error(f"❌ {p.name}: {e}")
            return False

    # Both URLs failed
    add_to_queue(filepath, event_type)
    return False

def _notify_deletion(filepath: str, state: dict):
    try:
        requests.post(
            f"http://{VESPER_IP}:{VESPER_PORT}/delete",
            json={"filepath": filepath},
            timeout=5
        )
        with _state_lock:
            state.pop(filepath, None)
            save_state(state, force=True)
        log.info(f"🗑️  {Path(filepath).name}")
    except Exception:
        pass

# ══════════════════════════════════════════
# COMMAND EXECUTION
# ══════════════════════════════════════════

def execute_command(command: str, args: dict) -> dict:
    try:
        if command == "create_file":
            path    = os.path.expanduser(args["path"])
            content = args.get("content", "")
            os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
            with open(path, "w") as f:
                f.write(content)
            return {"success": True, "result": f"Created: {path}"}

        elif command == "append_file":
            path = os.path.expanduser(args["path"])
            with open(path, "a") as f:
                f.write(f"\n\n{args.get('content', '')}")
            return {"success": True, "result": f"Appended: {path}"}

        elif command == "read_file":
            path = os.path.expanduser(args["path"])
            with open(path, "r", errors="ignore") as f:
                return {"success": True, "result": f.read()[:5000]}

        elif command == "delete_file":
            path = os.path.expanduser(args["path"])
            if os.path.exists(path):
                os.remove(path)
                return {"success": True, "result": f"Deleted: {path}"}
            return {"success": False, "result": "Not found"}

        elif command == "move_file":
            src = os.path.expanduser(args["src"])
            dst = os.path.expanduser(args["dst"])
            shutil.move(src, dst)
            return {"success": True, "result": f"Moved → {dst}"}

        elif command == "create_folder":
            path = os.path.expanduser(args["path"])
            os.makedirs(path, exist_ok=True)
            return {"success": True, "result": f"Created: {path}"}

        elif command == "delete_folder":
            path = os.path.expanduser(args["path"])
            shutil.rmtree(path, ignore_errors=True)
            return {"success": True, "result": f"Deleted: {path}"}

        elif command == "list_folder":
            path  = os.path.expanduser(args.get("path", "~/Documents"))
            items = []
            for item in sorted(os.listdir(path)):
                full = os.path.join(path, item)
                icon = "📁" if os.path.isdir(full) else "📄"
                try:
                    size = os.path.getsize(full)
                    items.append(f"{icon} {item} ({size//1024}KB)")
                except Exception:
                    items.append(f"{icon} {item}")
            return {"success": True, "result": "\n".join(items[:100])}

        elif command == "search_files":
            query   = args.get("query", "").lower()
            base    = os.path.expanduser(args.get("folder", "~"))
            matches = []
            for root, dirs, files in os.walk(base):
                dirs[:] = [d for d in dirs if d not in SKIP_DIRS]
                for f in files:
                    if query in f.lower():
                        matches.append(os.path.join(root, f))
                if len(matches) >= 50:
                    break
            return {"success": True, "result": "\n".join(matches)}

        elif command == "empty_trash":
            subprocess.run([
                "osascript", "-e",
                'tell application "Finder" to empty trash'
            ])
            return {"success": True, "result": "Trash emptied"}

        elif command == "move_to_trash":
            path = os.path.expanduser(args["path"])
            subprocess.run([
                "osascript", "-e",
                f'tell application "Finder" to delete POSIX file "{path}"'
            ])
            return {"success": True, "result": f"Trashed: {path}"}

        elif command == "open_file":
            subprocess.run(["open", os.path.expanduser(args["path"])])
            return {"success": True, "result": "Opened"}

        elif command == "open_app":
            subprocess.run(["open", "-a", args.get("app", "")])
            return {"success": True, "result": "App opened"}

        elif command == "notify":
            title   = args.get("title", "Vesper")
            message = args.get("message", "")
            subprocess.run(["osascript", "-e",
                f'display notification "{message}" with title "{title}"'
            ])
            return {"success": True, "result": "Notified"}

        elif command == "create_reminder":
            title = args.get("title", "")
            notes = args.get("notes", "")
            subprocess.run(["osascript", "-e",
                f'tell application "Reminders" to make new reminder '
                f'with properties {{name:"{title}", body:"{notes}"}}'
            ])
            return {"success": True, "result": f"Reminder: {title}"}

        elif command == "generate_pdf":
            path    = os.path.expanduser(args["path"])
            content = args.get("content", "")
            title   = args.get("title", "Document")
            try:
                from reportlab.lib.pagesizes import letter
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.platypus import (
                    SimpleDocTemplate, Paragraph, Spacer
                )
                from reportlab.lib.units import inch
                os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
                doc    = SimpleDocTemplate(path, pagesize=letter)
                styles = getSampleStyleSheet()
                story  = [
                    Paragraph(title, styles["Title"]),
                    Spacer(1, 0.3 * inch)
                ]
                for para in content.split("\n\n"):
                    if para.strip():
                        story.append(Paragraph(para.strip(), styles["Normal"]))
                        story.append(Spacer(1, 0.1 * inch))
                doc.build(story)
                subprocess.run(["open", path])
                return {"success": True, "result": f"PDF: {path}"}
            except ImportError:
                subprocess.run([
                    sys.executable, "-m", "pip", "install", "reportlab"
                ])
                return {"success": False, "result": "Installing reportlab, retry"}

        elif command == "run_script":
            script    = args.get("script", "")
            dangerous = ["rm -rf /", "sudo rm", "mkfs", "dd if"]
            if any(d in script for d in dangerous):
                return {"success": False, "result": "Blocked"}
            result = subprocess.run(
                script, shell=True,
                capture_output=True, text=True, timeout=30
            )
            return {
                "success": result.returncode == 0,
                "result": (result.stdout or result.stderr)[:2000]
            }

        return {"success": False, "result": f"Unknown: {command}"}

    except Exception as e:
        log.error(f"Command [{command}]: {e}")
        return {"success": False, "result": str(e)}

# ══════════════════════════════════════════
# COMMAND SERVER
# ══════════════════════════════════════════

class CommandHandler(http.server.BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            length  = int(self.headers.get("Content-Length", 0))
            data    = json.loads(self.rfile.read(length))
            result  = execute_command(
                data.get("command", ""),
                data.get("args", {})
            )
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps(result).encode())
        except Exception as e:
            self.send_response(500)
            self.end_headers()
            self.wfile.write(
                json.dumps({"success": False, "result": str(e)}).encode()
            )
    def log_message(self, *args):
        pass

def start_command_server():
    try:
        server = socketserver.TCPServer(
            ("0.0.0.0", COMMAND_PORT), CommandHandler
        )
        server.allow_reuse_address = True
        server.serve_forever()
    except Exception as e:
        log.error(f"Command server: {e}")

# ══════════════════════════════════════════
# QUEUE PROCESSOR
# ══════════════════════════════════════════

class QueueProcessor:
    def __init__(self, state: dict):
        self.state   = state
        self.running = True

    def run(self):
        while self.running:
            try:
                if is_mac_sleeping():
                    time.sleep(30)
                    continue
                if not can_reach_vesper():
                    time.sleep(15)
                    continue
                pending = load_queue()
                if not pending:
                    time.sleep(10)
                    continue
                log.info(f"📤 Flushing {len(pending)} queued")
                for item in pending.copy():
                    fp = item["filepath"]
                    et = item["event_type"]
                    # Remove from queue if: success, or file no longer exists,
                    # or content can't be extracted (permanently failing)
                    if sync_file(fp, et, self.state):
                        remove_from_queue(fp)
                    elif not os.path.exists(fp):
                        remove_from_queue(fp)
                    elif et != "deleted":
                        content = extract_content(fp)
                        if not content or len(content.strip()) < 20:
                            log.info(f"⏭️  Dropping unextractable: {Path(fp).name}")
                            remove_from_queue(fp)
                    time.sleep(SEND_DELAY)
            except Exception as e:
                log.error(f"Queue: {e}")
                time.sleep(30)

# ══════════════════════════════════════════
# FILE EVENT HANDLER
# ══════════════════════════════════════════

class VesperHandler(FileSystemEventHandler):
    def __init__(self, state: dict):
        self.state  = state
        self.timers = {}
        self.lock   = threading.Lock()

    def _debounce(self, filepath: str, event: str):
        with self.lock:
            if filepath in self.timers:
                self.timers[filepath].cancel()
            timer = threading.Timer(
                DEBOUNCE_SECS, self._process,
                args=[filepath, event]
            )
            self.timers[filepath] = timer
            timer.start()

    def _process(self, filepath: str, event: str):
        if not should_index(filepath):
            return
        if event != "deleted":
            if not os.path.exists(filepath):
                return
            with _state_lock:
                current = compute_hash(filepath)
                stored  = self.state.get(filepath, {}).get("hash")
            if current == stored:
                return
        if can_reach_vesper():
            sync_file(filepath, event, self.state)
        else:
            add_to_queue(filepath, event)

    def on_created(self, e):
        if not e.is_directory:
            self._debounce(e.src_path, "created")

    def on_modified(self, e):
        if not e.is_directory:
            self._debounce(e.src_path, "modified")

    def on_deleted(self, e):
        if not e.is_directory:
            self._process(e.src_path, "deleted")

    def on_moved(self, e):
        if not e.is_directory:
            self._process(e.src_path, "deleted")
            self._debounce(e.dest_path, "created")

# ══════════════════════════════════════════
# INITIAL SCAN — throttled
# ══════════════════════════════════════════

def count_files() -> int:
    count = 0
    for watch_dir in WATCH_DIRS:
        if not os.path.exists(watch_dir):
            continue
        for root, dirs, files in os.walk(watch_dir):
            dirs[:] = [
                d for d in dirs
                if d not in SKIP_DIRS and not d.startswith(".")
            ]
            for filename in files:
                if not filename.startswith("."):
                    filepath = os.path.join(root, filename)
                    if should_index(filepath):
                        count += 1
    return count

def initial_scan(state: dict):
    log.info("🔍 Counting files...")
    total = count_files()
    log.info(f"📁 Found {total} indexable files")
    progress.start(total)

    for watch_dir in WATCH_DIRS:
        if not os.path.exists(watch_dir):
            continue
        for root, dirs, files in os.walk(watch_dir):
            dirs[:] = [
                d for d in dirs
                if d not in SKIP_DIRS and not d.startswith(".")
            ]
            for filename in files:
                if filename.startswith("."):
                    continue
                filepath = os.path.join(root, filename)
                if not should_index(filepath):
                    continue

                with _state_lock:
                    current_hash = compute_hash(filepath)
                    stored_hash  = state.get(filepath, {}).get("hash")

                if current_hash == stored_hash:
                    progress.update(skipped=1)
                    continue

                is_new = filepath not in state
                event  = "created" if is_new else "modified"

                if can_reach_vesper():
                    success = sync_file(filepath, event, state)
                    if success:
                        progress.update(synced=1)
                        # Throttle between sends
                        time.sleep(SEND_DELAY)
                    else:
                        progress.update(failed=1)
                else:
                    add_to_queue(filepath, event)
                    progress.update(failed=1)

    # Force final state save
    with _state_lock:
        save_state(state, force=True)

    progress.done()

# ══════════════════════════════════════════
# AUTO-START
# ══════════════════════════════════════════

def install_autostart():
    plist_path = os.path.expanduser(
        "~/Library/LaunchAgents/com.vesper.filewatcher.plist"
    )
    if os.path.exists(plist_path):
        return

    python_path = sys.executable
    script_path = os.path.abspath(__file__)

    content = f"""<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE plist PUBLIC "-//Apple//DTD PLIST 1.0//EN"
  "http://www.apple.com/DTDs/PropertyList-1.0.dtd">
<plist version="1.0">
<dict>
    <key>Label</key>
    <string>com.vesper.filewatcher</string>
    <key>ProgramArguments</key>
    <array>
        <string>{python_path}</string>
        <string>{script_path}</string>
    </array>
    <key>RunAtLoad</key>
    <true/>
    <key>KeepAlive</key>
    <true/>
    <key>StandardOutPath</key>
    <string>{AGENT_DIR}/logs/watcher.log</string>
    <key>StandardErrorPath</key>
    <string>{AGENT_DIR}/logs/error.log</string>
</dict>
</plist>"""

    os.makedirs(os.path.dirname(plist_path), exist_ok=True)
    with open(plist_path, "w") as f:
        f.write(content)
    subprocess.run(["launchctl", "load", plist_path], capture_output=True)
    log.info("✅ Auto-start installed")

# ══════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════

def main():
    log.info("=" * 52)
    log.info("⚡ VESPER MAC AGENT - Production v6")
    log.info("=" * 52)
    log.info(f"Server:    {VESPER_IP}:{VESPER_PORT}")
    log.info(f"Throttle:  {SEND_DELAY}s between files")
    log.info(f"Watching:  {len(WATCH_DIRS)} directories")

    install_autostart()

    state = load_state()
    pending = load_queue()
    if pending:
        log.info(f"📦 {len(pending)} queued from last session")

    up = can_reach_vesper()
    log.info(f"Server:    {'✅ online' if up else '❌ offline'}")

    threading.Thread(
        target=start_command_server,
        daemon=True, name="commands"
    ).start()
    log.info(f"🎯 Commands on port {COMMAND_PORT}")

    qp = QueueProcessor(state)
    threading.Thread(target=qp.run, daemon=True, name="queue").start()
    log.info("📤 Queue processor started")

    threading.Thread(
        target=initial_scan, args=[state],
        daemon=True, name="scan"
    ).start()

    handler  = VesperHandler(state)
    observer = Observer()
    for d in WATCH_DIRS:
        observer.schedule(handler, d, recursive=True)
        log.info(f"👁️  {d}")

    observer.start()
    log.info("✅ Active — Vesper sees everything.\n")

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        log.info("Stopping...")
        qp.running = False
        observer.stop()

    observer.join()
    with _state_lock:
        save_state(state, force=True)
    log.info("Agent stopped.")

if __name__ == "__main__":
    main()